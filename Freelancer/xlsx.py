import csv
import openpyxl
from openpyxl import Workbook

# File path
excelFileName = 'example.xlsx'

# Opening the file as workbook
inputWorkbook = openpyxl.load_workbook(excelFileName)

# Load all the sheet names
sheets = inputWorkbook.sheetnames

outputWorkbook = Workbook()
outputSheet = outputWorkbook.active
#outputSheet = outputWorkbook.create_sheet('output')


outputRowList = []

for sh in sheets:
    imputSheet = inputWorkbook[sh]

    for i in range(1, imputSheet.max_row):
        singleRowList = []
        for j in range(1, imputSheet.max_column):
            eachCell = imputSheet.cell(row=i, column=j).value
            outputSheet.cell(row=i, column=j).value = eachCell
            singleRowList.append(eachCell)
        outputRowList.append(singleRowList)

#saving the output excel file
outputWorkbook.save('output.xlsx')

# Output CSV file
outputCSV = 'output.csv'

with open(outputCSV, 'w', newline='') as file:
    writer = csv.writer(file)
    writer.writerows(outputRowList)